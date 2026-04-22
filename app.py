import io
import re
import unicodedata
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import gspread
import pandas as pd
import streamlit as st
from google.oauth2.service_account import Credentials
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter


# =========================
# CONFIG
# =========================
DETAIL_SHEET_NAME = "No_consumieron"
SUMMARY_SHEET_NAME = "Resumen"
NOEL_SHEET_NAME = "Noel"
DATALAKE_SHEET_NAME = "Datalake"

MONTH_MAP = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
             "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"]

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.readonly",
]


# =========================
# UI
# =========================
st.set_page_config(page_title="Reservas - No consumieron", layout="wide")
st.title("Reservas - No consumieron")
st.caption("Sube el archivo de reservas y la app cruzará la información con Noel y Datalake desde Google Sheets.")


# =========================
# HELPERS
# =========================
def valor_texto(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def normalizar_documento(v) -> str:
    return re.sub(r"\D", "", valor_texto(v))


def normalizar_header(texto) -> str:
    texto = valor_texto(texto)
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")
    texto = re.sub(r"[\n\r\t]+", " ", texto)
    texto = re.sub(r"[^a-zA-Z0-9 ]", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip().lower()
    return texto


def limpiar_nombre_hoja(nombre: str) -> str:
    return re.sub(r'[\\\/\?\*\[\]\:]', "_", valor_texto(nombre))[:31]


def parse_fecha_flexible(valor) -> Optional[datetime]:
    if isinstance(valor, datetime):
        return valor

    if hasattr(valor, "to_pydatetime"):
        try:
            return valor.to_pydatetime()
        except Exception:
            pass

    texto = valor_texto(valor)
    if not texto:
        return None

    formatos = [
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%m/%d/%Y",
        "%d/%m/%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
    ]

    for fmt in formatos:
        try:
            return datetime.strptime(texto, fmt)
        except Exception:
            continue

    try:
        dt = pd.to_datetime(texto, errors="coerce", dayfirst=True)
        if pd.notna(dt):
            return dt.to_pydatetime()
    except Exception:
        pass

    return None


def obtener_periodo_desde_fecha(fecha: datetime) -> str:
    year = fecha.year
    month = fecha.month
    mes_texto = MONTH_MAP[month - 1]
    mes_num = f"{month:02d}"
    return f"{year}_{mes_num}_{mes_texto}"


def contar_unicos(arr: List[str]) -> int:
    return len({valor_texto(x) for x in arr if valor_texto(x)})


def clave_persona(r: dict) -> str:
    return normalizar_documento(r.get("cedula")) or f"{valor_texto(r.get('usuario'))}|{valor_texto(r.get('correo'))}"


def contar_personas_unicas(registros: List[dict]) -> int:
    return len({clave_persona(r) for r in registros})


def buscar_columna(headers: List[str], candidatos: List[str]) -> int:
    normalizados = [normalizar_header(h) for h in headers]

    for candidato in candidatos:
        exacto = normalizar_header(candidato)
        if exacto in normalizados:
            return normalizados.index(exacto)

    for candidato in candidatos:
        c = normalizar_header(candidato)
        for i, h in enumerate(normalizados):
            if c in h:
                return i

    raise ValueError(f"No encontré la columna requerida: {candidatos[0]}")


def buscar_columna_opcional(headers: List[str], candidatos: List[str]) -> int:
    normalizados = [normalizar_header(h) for h in headers]

    for candidato in candidatos:
        exacto = normalizar_header(candidato)
        if exacto in normalizados:
            return normalizados.index(exacto)

    for candidato in candidatos:
        c = normalizar_header(candidato)
        for i, h in enumerate(normalizados):
            if c in h:
                return i

    return -1


def obtener_columnas_reservas(headers: List[str]) -> dict:
    return {
        "fecha": buscar_columna(headers, ["fecha"]),
        "hora": buscar_columna(headers, ["hora"]),
        "numero": buscar_columna(headers, ["numero"]),
        "menu": buscar_columna(headers, ["menu"]),
        "usuario": buscar_columna(headers, ["usuario"]),
        "correo": buscar_columna(headers, ["correo electronico", "correo electrónico", "correo"]),
        "cedula": buscar_columna(headers, ["nit", "cedula", "cédula", "cc / nit"]),
        "matricula": buscar_columna_opcional(headers, ["matricula", "matrícula"]),
        "extension": buscar_columna_opcional(headers, ["extension", "extensión"]),
        "area": buscar_columna_opcional(headers, ["area", "área", "area reserva"]),
        "punto_venta": buscar_columna_opcional(headers, ["punto de venta"]),
        "lugar_entrega": buscar_columna_opcional(headers, ["lugar de entrega"]),
        "status_pedido": buscar_columna(headers, ["status del pedido"]),
    }


def obtener_columnas_noel(headers: List[str]) -> dict:
    return {
        "empresa": buscar_columna(headers, ["nombre de la empresa de acuerdo al nit", "empresa"]),
        "cedula": buscar_columna(headers, ["cedula", "cédula"]),
        "gerencia": buscar_columna_opcional(headers, ["gerencia"]),
        "nombre": 4,  # Columna E
    }


def obtener_columnas_datalake(headers: List[str]) -> dict:
    return {
        "descripcion": buscar_columna(headers, ["descripcion", "descripción", "empresa"]),
        "cedula": buscar_columna(headers, ["cedula", "cédula"]),
        "descripcion_gerencia": buscar_columna_opcional(headers, ["descripcion gerencia", "descripción gerencia", "gerencia"]),
        "nombre_jefe": buscar_columna_opcional(headers, ["nombre jefe", "jefe"]),
        "nombre": 6,      # Columna G
        "apellido1": 7,   # Columna H
        "apellido2": 8,   # Columna I
    }


def leer_excel_upload(file) -> pd.DataFrame:
    xls = pd.ExcelFile(file)
    first_valid_sheet = None

    for sheet_name in xls.sheet_names:
        file.seek(0)
        df = pd.read_excel(file, sheet_name=sheet_name)
        if df.shape[0] > 0 and df.shape[1] > 1:
            first_valid_sheet = sheet_name
            break

    file.seek(0)
    if first_valid_sheet is None:
        return pd.read_excel(file, sheet_name=0)

    return pd.read_excel(file, sheet_name=first_valid_sheet)


def leer_archivo_reservas(file) -> pd.DataFrame:
    nombre = file.name.lower()
    if nombre.endswith(".csv"):
        try:
            file.seek(0)
            return pd.read_csv(file)
        except Exception:
            file.seek(0)
            return pd.read_csv(file, sep=";")

    file.seek(0)
    return leer_excel_upload(file)


def es_empresa_noel(empresa: str) -> bool:
    e = normalizar_header(empresa)
    return e in {
        "compania de galletas noel s a s",
        "compania de galletas noel sas",
        "compania de galletas noel s a",
        "compania de galletas noel",
    }


# =========================
# GOOGLE SHEETS
# =========================
@st.cache_resource(show_spinner=False)
def get_gspread_client():
    secrets_dict = st.secrets.to_dict()

    if "gcp_service_account" not in secrets_dict:
        raise KeyError(
            'No encontré "gcp_service_account" en Secrets. '
            'Debes cargar la service account completa.'
        )

    creds_info = dict(st.secrets["gcp_service_account"])
    credentials = Credentials.from_service_account_info(creds_info, scopes=SCOPES)
    return gspread.authorize(credentials)


def obtener_master_sheet_url() -> str:
    secrets_dict = st.secrets.to_dict()

    master_url = secrets_dict.get("MASTER_SHEET_URL")

    if not master_url and "gcp_service_account" in secrets_dict:
        gcp_block = secrets_dict.get("gcp_service_account", {})
        if isinstance(gcp_block, dict):
            master_url = gcp_block.get("MASTER_SHEET_URL")

    if not master_url:
        raise KeyError(
            f'No encontré "MASTER_SHEET_URL" en Secrets. '
            f'Claves disponibles: {list(secrets_dict.keys())}'
        )

    return master_url


def leer_google_sheet(url: str, worksheet_name: str) -> pd.DataFrame:
    gc = get_gspread_client()
    sh = gc.open_by_url(url)

    disponibles = [ws.title for ws in sh.worksheets()]
    if worksheet_name not in disponibles:
        raise ValueError(
            f'No encontré la pestaña "{worksheet_name}". '
            f'Pestañas disponibles: {", ".join(disponibles)}'
        )

    ws = sh.worksheet(worksheet_name)
    records = ws.get_all_records()
    return pd.DataFrame(records)


def cargar_bases_maestras() -> Tuple[pd.DataFrame, pd.DataFrame]:
    master_url = obtener_master_sheet_url()
    df_noel = leer_google_sheet(master_url, NOEL_SHEET_NAME)
    df_datalake = leer_google_sheet(master_url, DATALAKE_SHEET_NAME)
    return df_noel, df_datalake


def construir_indice_noel(df: pd.DataFrame) -> Dict[str, dict]:
    if df.empty:
        return {}

    headers = list(df.columns)
    col = obtener_columnas_noel(headers)
    rows = df.fillna("").values.tolist()
    idx = {}

    for row in rows:
        cedula = normalizar_documento(row[col["cedula"]])
        if not cedula:
            continue

        nombre_completo = valor_texto(row[col["nombre"]]) if len(row) > col["nombre"] else ""

        idx[cedula] = {
            "empresa": valor_texto(row[col["empresa"]]),
            "gerencia": valor_texto(row[col["gerencia"]]) if col["gerencia"] != -1 else "",
            "jefe": "",
            "nombreCompleto": nombre_completo,
            "fuenteCruce": "Noel",
            "encontrado": True,
        }

    return idx


def construir_indice_datalake(df: pd.DataFrame) -> Dict[str, dict]:
    if df.empty:
        return {}

    headers = list(df.columns)
    col = obtener_columnas_datalake(headers)
    rows = df.fillna("").values.tolist()
    idx = {}

    for row in rows:
        cedula = normalizar_documento(row[col["cedula"]])
        if not cedula:
            continue

        nombre = valor_texto(row[col["nombre"]]) if len(row) > col["nombre"] else ""
        apellido1 = valor_texto(row[col["apellido1"]]) if len(row) > col["apellido1"] else ""
        apellido2 = valor_texto(row[col["apellido2"]]) if len(row) > col["apellido2"] else ""
        nombre_completo = " ".join(x for x in [nombre, apellido1, apellido2] if x).strip()

        idx[cedula] = {
            "empresa": valor_texto(row[col["descripcion"]]),
            "gerencia": valor_texto(row[col["descripcion_gerencia"]]) if col["descripcion_gerencia"] != -1 else "",
            "jefe": valor_texto(row[col["nombre_jefe"]]) if col["nombre_jefe"] != -1 else "",
            "nombreCompleto": nombre_completo,
            "fuenteCruce": "Datalake",
            "encontrado": True,
        }

    return idx


# =========================
# RESÚMENES
# =========================
def construir_top_usuarios(registros: List[dict]) -> List[List]:
    mapa = {}

    for r in registros:
        key = clave_persona(r)
        if key not in mapa:
            mapa[key] = {
                "usuario": r["usuario"],
                "cedula": r["cedula"],
                "empresa": r["empresa"],
                "cantidad": 0,
            }
        mapa[key]["cantidad"] += 1

    return [
        [x["usuario"], x["cedula"], x["empresa"], x["cantidad"]]
        for x in sorted(mapa.values(), key=lambda x: (-x["cantidad"], x["usuario"]))
    ]


def construir_top_usuarios_detallado(registros: List[dict]) -> List[List]:
    mapa = {}

    for r in registros:
        key = clave_persona(r)
        if key not in mapa:
            mapa[key] = {
                "usuario": r["usuario"],
                "cedula": r["cedula"],
                "empresa": r["empresa"],
                "gerencia": r["gerencia"],
                "cantidad": 0,
            }
        mapa[key]["cantidad"] += 1

    return [
        [x["usuario"], x["cedula"], x["empresa"], x["gerencia"], x["cantidad"]]
        for x in sorted(mapa.values(), key=lambda x: (-x["cantidad"], x["usuario"]))
    ]


def resumir_por_empresa(registros: List[dict]) -> List[List]:
    mapa = {}

    for r in registros:
        empresa = valor_texto(r["empresa"]) or "SIN CRUCE"
        if empresa not in mapa:
            mapa[empresa] = {
                "reservas": 0,
                "personas": set(),
                "gerencias": set(),
            }

        mapa[empresa]["reservas"] += 1
        mapa[empresa]["personas"].add(clave_persona(r))
        if valor_texto(r["gerencia"]):
            mapa[empresa]["gerencias"].add(r["gerencia"])

    data = []
    for empresa, v in mapa.items():
        data.append([
            empresa,
            v["reservas"],
            len(v["personas"]),
            ", ".join(sorted(v["gerencias"]))
        ])

    return sorted(data, key=lambda x: (-x[1], str(x[0])))


def resumir_por_gerencia(registros: List[dict]) -> List[List]:
    mapa = {}

    for r in registros:
        gerencia = valor_texto(r["gerencia"]) or "SIN CRUCE"
        if gerencia not in mapa:
            mapa[gerencia] = {
                "reservas": 0,
                "personas": set(),
                "empresas": set(),
            }

        mapa[gerencia]["reservas"] += 1
        mapa[gerencia]["personas"].add(clave_persona(r))
        if valor_texto(r["empresa"]):
            mapa[gerencia]["empresas"].add(r["empresa"])

    data = []
    for gerencia, v in mapa.items():
        data.append([
            gerencia,
            v["reservas"],
            len(v["personas"]),
            ", ".join(sorted(v["empresas"]))
        ])

    return sorted(data, key=lambda x: (-x[1], str(x[0])))


def resumir_por_jefe(registros: List[dict]) -> List[List]:
    mapa = {}

    for r in registros:
        jefe = valor_texto(r["jefe"]) or "SIN CRUCE"
        if jefe not in mapa:
            mapa[jefe] = {
                "reservas": 0,
                "personas": set(),
                "empresas": set(),
            }

        mapa[jefe]["reservas"] += 1
        mapa[jefe]["personas"].add(clave_persona(r))
        if valor_texto(r["empresa"]):
            mapa[jefe]["empresas"].add(r["empresa"])

    data = []
    for jefe, v in mapa.items():
        data.append([
            jefe,
            v["reservas"],
            len(v["personas"]),
            ", ".join(sorted(v["empresas"]))
        ])

    return sorted(data, key=lambda x: (-x[1], str(x[0])))


def construir_sin_cruce(registros: List[dict]) -> List[List]:
    return [
        [r["usuario"], r["cedula"], r["fecha"], r["areaReserva"]]
        for r in registros
        if not r["encontradoCruce"]
    ]


# =========================
# MÉTRICAS
# =========================
def construir_metricas_desde_archivo(df_reservas: pd.DataFrame, col: dict) -> dict:
    rows = df_reservas.fillna("").to_dict("records")

    total_reservas = len(rows)
    total_consumieron = 0
    total_no_consumieron = 0
    personas_unicas_no_consumieron = set()

    for row in rows:
        registro = {
            "cedula": row[df_reservas.columns[col["cedula"]]] if col["cedula"] != -1 else "",
            "usuario": valor_texto(row[df_reservas.columns[col["usuario"]]]) if col["usuario"] != -1 else "",
            "correo": valor_texto(row[df_reservas.columns[col["correo"]]]) if col["correo"] != -1 else "",
        }

        key = clave_persona(registro)
        estado = valor_texto(row[df_reservas.columns[col["status_pedido"]]]).lower()

        if estado == "delivered":
            total_consumieron += 1
        elif estado == "accepted":
            total_no_consumieron += 1
            if key:
                personas_unicas_no_consumieron.add(key)

    return {
        "totalReservas": total_reservas,
        "personasConsumieron": total_consumieron,
        "personasNoConsumieron": total_no_consumieron,
        "personasUnicas": len(personas_unicas_no_consumieron),
    }


# =========================
# PROCESAMIENTO
# =========================
def procesar_reservas(df_reservas: pd.DataFrame, df_noel: pd.DataFrame, df_datalake: pd.DataFrame) -> Tuple[List[dict], dict]:
    if df_reservas.empty:
        raise ValueError("El archivo de reservas no tiene datos.")

    df_reservas = df_reservas.copy()
    df_reservas.columns = [valor_texto(c) for c in df_reservas.columns]
    headers = list(df_reservas.columns)
    col = obtener_columnas_reservas(headers)

    indice_noel = construir_indice_noel(df_noel)
    indice_datalake = construir_indice_datalake(df_datalake)

    metricas = construir_metricas_desde_archivo(df_reservas, col)

    rows = df_reservas.fillna("").values.tolist()

    filtrados = []
    for row in rows:
        estado = valor_texto(row[col["status_pedido"]]).lower()
        if estado == "accepted":
            filtrados.append(row)

    registros = []
    for row in filtrados:
        cedula_original = valor_texto(row[col["cedula"]])
        cedula_normalizada = normalizar_documento(cedula_original)

        cruce_noel = indice_noel.get(cedula_normalizada)
        cruce_datalake = indice_datalake.get(cedula_normalizada)

        cruce = cruce_noel or cruce_datalake or {}

        empresa = valor_texto(cruce.get("empresa"))
        gerencia = valor_texto(cruce.get("gerencia"))
        jefe = valor_texto(cruce.get("jefe"))
        fuente_cruce = valor_texto(cruce.get("fuenteCruce"))
        nombre_completo = valor_texto(cruce.get("nombreCompleto"))

        if es_empresa_noel(empresa) and cruce_datalake:
            if not jefe:
                jefe = valor_texto(cruce_datalake.get("jefe"))
            if not nombre_completo:
                nombre_completo = valor_texto(cruce_datalake.get("nombreCompleto"))

        usuario_reserva = valor_texto(row[col["usuario"]])
        usuario_final = nombre_completo or usuario_reserva

        registros.append({
            "fecha": row[col["fecha"]],
            "hora": row[col["hora"]],
            "numero": row[col["numero"]],
            "menu": row[col["menu"]],
            "usuario": usuario_final,
            "correo": valor_texto(row[col["correo"]]),
            "cedula": cedula_original,
            "matricula": valor_texto(row[col["matricula"]]) if col["matricula"] != -1 else "",
            "extension": valor_texto(row[col["extension"]]) if col["extension"] != -1 else "",
            "areaReserva": valor_texto(row[col["area"]]) if col["area"] != -1 else "",
            "puntoVenta": valor_texto(row[col["punto_venta"]]) if col["punto_venta"] != -1 else "",
            "lugarEntrega": valor_texto(row[col["lugar_entrega"]]) if col["lugar_entrega"] != -1 else "",
            "statusPedido": valor_texto(row[col["status_pedido"]]),
            "empresa": empresa,
            "gerencia": gerencia,
            "jefe": jefe,
            "fuenteCruce": fuente_cruce,
            "encontradoCruce": bool(cruce.get("encontrado", False)),
        })

    resultado = {
        "total": len(registros),
        "personasUnicasNoConsumieron": contar_personas_unicas(registros),
        **metricas,
    }

    return registros, resultado


# =========================
# DATAFRAMES DE SALIDA
# =========================
def df_usuarios(registros: List[dict]) -> pd.DataFrame:
    return pd.DataFrame(
        construir_top_usuarios(registros),
        columns=["Usuario", "CC / Nit", "Empresa", "Cantidad"]
    )


def df_detalle(registros: List[dict]) -> pd.DataFrame:
    return pd.DataFrame([{
        "Fecha": r["fecha"],
        "Hora": r["hora"],
        "Numero": r["numero"],
        "Menu": r["menu"],
        "Usuario": r["usuario"],
        "Correo electrónico": r["correo"],
        "CC / Nit": r["cedula"],
        "Empresa": r["empresa"],
        "Gerencia": r["gerencia"],
        "Jefe": r["jefe"],
        "Fuente cruce": r["fuenteCruce"],
        "Matricula": r["matricula"],
        "Extensión": r["extension"],
        "Área reserva": r["areaReserva"],
        "Punto de venta": r["puntoVenta"],
        "Lugar de entrega": r["lugarEntrega"],
        "Status del pedido": r["statusPedido"],
    } for r in registros])


# =========================
# EXCEL
# =========================
FILL_BLUE = PatternFill(fill_type="solid", fgColor="D9EAF7")
FILL_DARK = PatternFill(fill_type="solid", fgColor="1F4E78")
FILL_SECTION = PatternFill(fill_type="solid", fgColor="EAF2F8")
FILL_HEADER = PatternFill(fill_type="solid", fgColor="F4F6F6")
FONT_BOLD = Font(bold=True)
FONT_TITLE = Font(bold=True, size=14, color="FFFFFF")


def auto_fit_ws(ws):
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 45)


def formatear_hoja_base_excel(ws, total_columns: int):
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = FONT_BOLD
        cell.fill = FILL_BLUE

    last_row = max(ws.max_row, 1)
    ws.auto_filter = AutoFilter(ref=f"A1:{get_column_letter(total_columns)}{last_row}")
    auto_fit_ws(ws)


def formatear_resumen_excel(ws):
    for cell in ws["1:1"]:
        if cell.column <= 2 or cell.column >= 4:
            cell.font = FONT_BOLD
            cell.fill = FILL_BLUE

    for cell in ws["2:2"]:
        if cell.column >= 4:
            cell.font = FONT_BOLD
            cell.fill = FILL_BLUE

    ws.freeze_panes = "A2"
    auto_fit_ws(ws)


def escribir_tabla_resumen_excel(ws, start_row: int, titulo: str, headers: List[str], values: List[List]) -> int:
    ws.cell(start_row, 1, titulo)
    for j, h in enumerate(headers, start=1):
        ws.cell(start_row + 1, j, h)

    if values:
        for i, row in enumerate(values, start=start_row + 2):
            for j, val in enumerate(row, start=1):
                ws.cell(i, j, val)
        return start_row + 2 + len(values)

    ws.cell(start_row + 2, 1, "Sin datos")
    return start_row + 3


def construir_excel(registros: List[dict]) -> bytes:
    wb = Workbook()

    ws_detalle = wb.active
    ws_detalle.title = DETAIL_SHEET_NAME

    headers_detalle = [
        "Fecha", "Hora", "Numero", "Menu", "Usuario", "Correo electrónico",
        "CC / Nit", "Empresa", "Gerencia", "Jefe", "Fuente cruce",
        "Matricula", "Extensión", "Área reserva", "Punto de venta",
        "Lugar de entrega", "Status del pedido"
    ]
    ws_detalle.append(headers_detalle)

    for r in registros:
        ws_detalle.append([
            r["fecha"], r["hora"], r["numero"], r["menu"], r["usuario"], r["correo"],
            r["cedula"], r["empresa"], r["gerencia"], r["jefe"], r["fuenteCruce"],
            r["matricula"], r["extension"], r["areaReserva"], r["puntoVenta"],
            r["lugarEntrega"], r["statusPedido"]
        ])

    formatear_hoja_base_excel(ws_detalle, len(headers_detalle))

    ws_resumen = wb.create_sheet(SUMMARY_SHEET_NAME)
    ws_resumen["A1"] = "Indicador"
    ws_resumen["B1"] = "Valor"
    ws_resumen["A2"] = "Total no consumieron"
    ws_resumen["B2"] = len(registros)
    ws_resumen["A3"] = "Personas únicas"
    ws_resumen["B3"] = contar_personas_unicas(registros)

    ws_resumen["D1"] = "Top usuarios"
    resumen_top_headers = ["Usuario", "CC / Nit", "Empresa", "Cantidad"]
    for j, h in enumerate(resumen_top_headers, start=4):
        ws_resumen.cell(2, j, h)

    top_usuarios = construir_top_usuarios(registros)
    for i, row in enumerate(top_usuarios, start=3):
        for j, val in enumerate(row, start=4):
            ws_resumen.cell(i, j, val)

    formatear_resumen_excel(ws_resumen)

    fecha_base = None
    for r in registros:
        fecha_base = parse_fecha_flexible(r["fecha"])
        if fecha_base:
            break

    if not fecha_base:
        fecha_base = datetime.now()

    periodo = obtener_periodo_desde_fecha(fecha_base)
    mes = periodo.split("_")[2]
    nombre_informe = limpiar_nombre_hoja(f"INF_{mes}")
    ws_inf = wb.create_sheet(nombre_informe)

    total = len(registros)
    personas_unicas = contar_personas_unicas(registros)
    empresas_unicas = contar_unicos([r["empresa"] for r in registros])
    gerencias_unicas = contar_unicos([r["gerencia"] for r in registros])
    jefes_unicos = contar_unicos([r["jefe"] for r in registros])

    ws_inf["A1"] = f"INFORME NO CONSUMIERON - {periodo}"
    ws_inf.merge_cells("A1:G1")
    ws_inf["A1"].font = FONT_TITLE
    ws_inf["A1"].fill = FILL_DARK

    ws_inf["F3"] = "Generado el"
    ws_inf["G3"] = datetime.now()

    ws_inf["A5"] = "KPI"
    ws_inf["B5"] = "Valor"
    ws_inf["A6"] = "Total no consumieron"
    ws_inf["B6"] = total
    ws_inf["A7"] = "Personas únicas"
    ws_inf["B7"] = personas_unicas
    ws_inf["A8"] = "Empresas únicas"
    ws_inf["B8"] = empresas_unicas
    ws_inf["A9"] = "Gerencias únicas"
    ws_inf["B9"] = gerencias_unicas
    ws_inf["A10"] = "Jefes únicos"
    ws_inf["B10"] = jefes_unicos

    for cell in ws_inf["5:5"]:
        if cell.column <= 2:
            cell.font = FONT_BOLD
            cell.fill = FILL_BLUE

    row = 13

    row = escribir_tabla_resumen_excel(
        ws_inf, row,
        "Resumen por empresa",
        ["Empresa", "Reservas no consumidas", "Personas únicas", "Gerencias"],
        resumir_por_empresa(registros)
    )

    row += 2
    row = escribir_tabla_resumen_excel(
        ws_inf, row,
        "Resumen por gerencia",
        ["Gerencia", "Reservas no consumidas", "Personas únicas", "Empresas"],
        resumir_por_gerencia(registros)
    )

    row += 2
    row = escribir_tabla_resumen_excel(
        ws_inf, row,
        "Resumen por jefe",
        ["Jefe", "Reservas no consumidas", "Personas únicas", "Empresas"],
        resumir_por_jefe(registros)
    )

    row += 2
    row = escribir_tabla_resumen_excel(
        ws_inf, row,
        "Top personas reincidentes",
        ["Usuario", "CC / Nit", "Empresa", "Gerencia", "Cantidad"],
        construir_top_usuarios_detallado(registros)
    )

    row += 2
    row = escribir_tabla_resumen_excel(
        ws_inf, row,
        "Sin cruce",
        ["Usuario", "CC / Nit", "Fecha", "Área reserva"],
        construir_sin_cruce(registros)
    )

    row += 2
    row = escribir_tabla_resumen_excel(
        ws_inf, row,
        "Detalle completo",
        [
            "Fecha", "Hora", "Numero", "Usuario", "CC / Nit", "Empresa",
            "Gerencia", "Jefe", "Área reserva", "Menu", "Fuente cruce",
            "Status del pedido"
        ],
        [
            [
                r["fecha"], r["hora"], r["numero"], r["usuario"], r["cedula"],
                r["empresa"], r["gerencia"], r["jefe"], r["areaReserva"],
                r["menu"], r["fuenteCruce"], r["statusPedido"]
            ]
            for r in registros
        ]
    )

    for r in range(13, ws_inf.max_row + 1):
        titulo = valor_texto(ws_inf.cell(r, 1).value)
        siguiente = ws_inf.cell(r + 1, 1).value if r + 1 <= ws_inf.max_row else None
        if titulo and siguiente:
            for c in range(1, ws_inf.max_column + 1):
                ws_inf.cell(r, c).font = FONT_BOLD
                ws_inf.cell(r, c).fill = FILL_SECTION
                ws_inf.cell(r + 1, c).font = FONT_BOLD
                ws_inf.cell(r + 1, c).fill = FILL_HEADER

    ws_inf.freeze_panes = "A5"
    auto_fit_ws(ws_inf)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# =========================
# MAIN
# =========================
uploaded_file = st.file_uploader(
    "Sube el archivo de reservas (.xlsx, .xls o .csv)",
    type=["xlsx", "xls", "csv"]
)

if uploaded_file:
    try:
        with st.spinner("Leyendo bases maestras desde Google Sheets..."):
            df_noel, df_datalake = cargar_bases_maestras()

        with st.spinner("Procesando archivo de reservas..."):
            df_reservas = leer_archivo_reservas(uploaded_file)
            registros, resultado = procesar_reservas(df_reservas, df_noel, df_datalake)

        usuarios = df_usuarios(registros)

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total reservas", resultado["totalReservas"])
        c2.metric("Personas consumieron", resultado["personasConsumieron"])
        c3.metric("Personas no consumieron", resultado["personasNoConsumieron"])
        c4.metric("Personas únicas", resultado["personasUnicas"])

        st.subheader("USUARIOS")
        st.dataframe(usuarios, use_container_width=True, height=450)

        excel_bytes = construir_excel(registros)
        nombre_salida = f"resultado_no_consumieron_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        st.download_button(
            "Descargar Excel resultado",
            data=excel_bytes,
            file_name=nombre_salida,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.success("Proceso completado correctamente.")

    except Exception as e:
        st.error(f"Error: {e}")
